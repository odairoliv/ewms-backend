import io
from decimal import Decimal, InvalidOperation

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.database import SessionLocal
from app.models import Apontamento, Upload
from app.repositories.apontamento_repository import ApontamentoRepository
from app.repositories.consultor_repository import ConsultorRepository
from app.repositories.custom_field_repository import CustomFieldRepository
from app.repositories.upload_repository import UploadRepository
from app.services.id_generator import next_id
from app.services.periodo_utils import calcular_semana, periodo_atual, periodo_para_data

# Colunas fixas do template — o nome de exibição e a obrigatoriedade podem ser
# sobrescritos por empresa em tb_custom (id_objeto="apontamento"), igual aos
# formulários/tabelas. id_consultor/id_fornecedor/periodo/horas_trabalhadas são
# obrigatórios por padrão; id_projeto/nome_projeto são opcionais por padrão.
COLUNAS_BASE = [
    {"nome_coluna": "id_consultor", "nome": "ID Consultor", "obrigatorio": True, "tipo": "texto"},
    {"nome_coluna": "id_fornecedor", "nome": "ID Fornecedor", "obrigatorio": True, "tipo": "texto"},
    {"nome_coluna": "id_projeto", "nome": "ID Projeto", "obrigatorio": False, "tipo": "texto"},
    {"nome_coluna": "nome_projeto", "nome": "Nome do Projeto", "obrigatorio": False, "tipo": "texto"},
    {"nome_coluna": "periodo", "nome": "Período (Mês/Ano)", "obrigatorio": True, "tipo": "texto"},
    {"nome_coluna": "horas_trabalhadas", "nome": "Horas Trabalhadas", "obrigatorio": True, "tipo": "float"},
]

CUSTOM_TIPOS_FIXOS = {
    "custom01": "date",
    "custom02": "texto",
    "custom03": "texto",
    "custom04": "float",
}


class UploadService:
    def __init__(self, db: Session):
        self.db = db
        self.repo = UploadRepository(db)
        self.apontamento_repo = ApontamentoRepository(db)
        self.consultor_repo = ConsultorRepository(db)
        self.custom_field_repo = CustomFieldRepository(db)

    def list_for_user(self, current_user: dict) -> list[Upload]:
        if current_user.get("perfil") == "SuperUsuario":
            return self.repo.list_all()
        return self.repo.list_by_empresa(current_user.get("empresa_id"))

    def template_colunas(self, empresa_id: str) -> list[dict]:
        """Monta a lista de colunas do template de upload para a empresa: campos base
        (com nome/obrigatoriedade que a empresa pode ter customizado) + Custom01-04
        visíveis para o objeto "apontamento" daquela empresa."""
        configs = self.custom_field_repo.list_by_empresa_objeto(empresa_id, "apontamento")
        colunas = []
        for base in COLUNAS_BASE:
            cfg = next((c for c in configs if c.nome_coluna == base["nome_coluna"]), None)
            colunas.append({
                "nome_coluna": base["nome_coluna"],
                "nome": cfg.nome if cfg else base["nome"],
                "obrigatorio": cfg.obrigatorio if cfg else base["obrigatorio"],
                "tipo": base["tipo"],
            })
        for nome_coluna, tipo in CUSTOM_TIPOS_FIXOS.items():
            cfg = next((c for c in configs if c.nome_coluna == nome_coluna), None)
            if cfg and cfg.visivel:
                colunas.append({
                    "nome_coluna": nome_coluna, "nome": cfg.nome, "obrigatorio": cfg.obrigatorio, "tipo": tipo,
                })
        return colunas

    def iniciar_upload(self, filename: str, periodo: str | None, current_user: dict) -> Upload:
        """Cria o registro de upload com status "Processando" e devolve na hora —
        o processamento pesado roda em background (ver processar_em_background)."""
        if not filename.endswith(".xlsx"):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Arquivo deve ser .xlsx")
        upload = Upload(
            id=next_id("UP", self.repo.count()),
            empresa_id=current_user.get("empresa_id"),
            arquivo=filename,
            periodo=periodo or periodo_atual(),
            status="Processando",
        )
        return self.repo.create(upload)

    def _validar_linha(self, colunas: list[dict], cabecalho_idx: dict, linha: tuple, num_linha: int, current_user: dict) -> dict:
        """Valida uma linha de dados contra as colunas do template. Levanta ValueError
        com mensagem indicando linha/coluna/motivo na primeira inconsistência encontrada."""
        is_super = current_user.get("perfil") == "SuperUsuario"
        empresa_id = current_user.get("empresa_id")
        dados: dict = {}

        for coluna in colunas:
            idx = cabecalho_idx.get(coluna["nome_coluna"])
            valor = linha[idx] if idx is not None and idx < len(linha) else None

            if coluna["obrigatorio"] and (valor is None or str(valor).strip() == ""):
                raise ValueError(
                    f"Linha {num_linha}, coluna '{coluna['nome']}': valor obrigatório não informado"
                )
            if valor is None or str(valor).strip() == "":
                dados[coluna["nome_coluna"]] = None
                continue

            if coluna["tipo"] == "float":
                try:
                    dados[coluna["nome_coluna"]] = Decimal(str(valor))
                except InvalidOperation:
                    raise ValueError(
                        f"Linha {num_linha}, coluna '{coluna['nome']}': valor '{valor}' não é um número válido"
                    )
            else:
                dados[coluna["nome_coluna"]] = str(valor).strip()

        try:
            periodo_para_data(str(dados["periodo"]))
        except ValueError:
            raise ValueError(
                f"Linha {num_linha}, coluna 'Período': '{dados['periodo']}' não está no formato Mês/Ano (ex: Maio/2026)"
            )

        consultor = self.consultor_repo.get_by_id(dados["id_consultor"])
        if not consultor:
            raise ValueError(
                f"Linha {num_linha}, coluna 'ID Consultor': consultor '{dados['id_consultor']}' não encontrado"
            )
        if not is_super and (not consultor.fornecedor or consultor.fornecedor.empresa_id != empresa_id):
            raise ValueError(
                f"Linha {num_linha}, coluna 'ID Consultor': consultor '{dados['id_consultor']}' não pertence à sua empresa"
            )

        dados["_consultor"] = consultor
        return dados

    @staticmethod
    def processar_em_background(upload_id: str, conteudo: bytes, current_user: dict) -> None:
        """Roda fora do ciclo da requisição (FastAPI BackgroundTasks), com sua própria
        sessão de banco — o usuário pode navegar para outra tela enquanto isso acontece."""
        db = SessionLocal()
        try:
            service = UploadService(db)
            upload = service.repo.get_by_id(upload_id)
            if not upload:
                return

            total_linhas = 0
            try:
                colunas = service.template_colunas(upload.empresa_id)
                workbook = load_workbook(filename=io.BytesIO(conteudo), data_only=True)
                sheet = workbook.active

                cabecalho = [str(c).strip() if c else None for c in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
                cabecalho_idx = {}
                for coluna in colunas:
                    for idx, titulo in enumerate(cabecalho):
                        if titulo and titulo.strip().lower() == coluna["nome"].strip().lower():
                            cabecalho_idx[coluna["nome_coluna"]] = idx
                            break

                linhas = [l for l in sheet.iter_rows(min_row=2, values_only=True) if l and l[0] is not None]
                total_linhas = len(linhas)

                dados_validados = []
                for i, linha in enumerate(linhas, start=2):
                    dados_validados.append(service._validar_linha(colunas, cabecalho_idx, linha, i, current_user))

                # Tudo validado — agora persiste tudo de uma vez.
                for dados in dados_validados:
                    consultor = dados.pop("_consultor")
                    horas_decimal = dados["horas_trabalhadas"]
                    valor_total = (horas_decimal * consultor.valor_hora).quantize(Decimal("0.01"))
                    data_reporte = periodo_para_data(str(dados["periodo"]))
                    apontamento = Apontamento(
                        id=next_id("APO", service.apontamento_repo.count()),
                        id_consultor=consultor.id,
                        id_fornecedor=dados.get("id_fornecedor") or consultor.id_fornecedor,
                        id_projeto=dados.get("id_projeto"),
                        nome_projeto=dados.get("nome_projeto"),
                        data_reporte=data_reporte,
                        semana=calcular_semana(data_reporte),
                        horas_trabalhadas=horas_decimal,
                        valor_hora=consultor.valor_hora,
                        valor_total=valor_total,
                        status="Em Análise",
                        custom01=dados.get("custom01"),
                        custom02=dados.get("custom02"),
                        custom03=dados.get("custom03"),
                        custom04=dados.get("custom04"),
                    )
                    service.apontamento_repo.create(apontamento)

                upload.linhas = len(linhas)
                upload.processadas = len(dados_validados)
                upload.rejeitadas = 0
                upload.status = "Processado"
                upload.erro_detalhe = None
                service.repo.update(upload)

            except ValueError as e:
                upload.linhas = total_linhas
                upload.processadas = 0
                upload.rejeitadas = total_linhas
                upload.status = "Recusado"
                upload.erro_detalhe = str(e)
                service.repo.update(upload)
            except Exception as e:  # noqa: BLE001 - upload não pode deixar o background "travado" sem status final
                upload.status = "Recusado"
                upload.erro_detalhe = f"Erro inesperado ao processar o arquivo: {e}"
                service.repo.update(upload)
        finally:
            db.close()
